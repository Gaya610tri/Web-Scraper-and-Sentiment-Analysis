
# Import the dependency

import openpyxl
import requests
from bs4 import BeautifulSoup


# load the excel file Input.xlsx

wrkbk = openpyxl.load_workbook("Input.xlsx") 

sh = wrkbk.active 

# Create a function to extract data from the web page with urls given in the Input.xlsx file

def dataExtract(url_id, url):
    # Send an HTTP request to the URL of the webpage you want to access
    response = requests.get(url)
    if response.status_code == 200:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, "html.parser")
        #get title and content of webpage
        #<div class="td-post-content tagdiv-type">
        title = soup.find('h1').get_text()
        content = soup.find('div', class_ = 'td-post-content').get_text()

        #write to the file
        fileName = url_id + '.txt'
        with open(fileName, 'w', encoding="utf-8") as file:
            file.write(title)
            file.write('\n')
            file.write(content)
            file.close()

# extract data for all the rows in the file

for i in range(2, sh.max_row+1): 
#     print("\n") 
#     print("Row ", i, " data :") 
    cell_obj = sh.cell(row=i, column=1)
    url_id = cell_obj.value
    cell_obj = sh.cell(row=i, column=2)
    url = cell_obj.value
    dataExtract(url_id, url)


# Preprocess the data by removing punctuations and splitting the content into list of tokens

def preProcess(text):
    text = text.lower()
    skips = ['!', "'", '"', "#", "$", "%", "&", "()", "*", "+", ",", "-", ".", "/", ":", ";", "<", "=", ">", "?", "@", "[", "]", "^", "_","`", "{", "|", "}", "~", "\n"]
    for ch in skips:  
        text = text.replace(ch, "") 
        
    tokens = text.split(" ")
    return tokens

# Create stop words list from the files provided

def createStopWords(files):
    totalData = ""
    for fileName in files:
        with open(fileName, 'r', encoding='latin-1') as file:
              data = file.read().replace('\n', ' ')
        totalData += " " + data
        
    StopWords = preProcess(totalData)
    return StopWords

files = ['StopWords_Auditor.txt', 'StopWords_Currencies.txt', 'StopWords_DatesandNumbers.txt', 'StopWords_Generic.txt', 'StopWords_GenericLong.txt', 'StopWords_Geographic.txt', 'StopWords_Names.txt']
StopWords = createStopWords(files)


# Remove the stop words from the tokens

def filterContent(content):
    processedContent = preProcess(content)
    filtered_tokens = []
    for word in processedContent:
        if(word not in StopWords):
            filtered_tokens.append(word)
    return filtered_tokens

# #### SENTIMENT ANALYSIS

# retrieve positive words from positive-words.txt file

positiveWords = []
with open('positive-words.txt', 'r') as file1:
              data = file1.read().replace('\n', ' ')
positiveWords = preProcess(data)

# POSITIVE SCORE

def calculatePosScore(filtered_tokens):
    positiveScore = 0
    for token in filtered_tokens:
        if(token in positiveWords):
            positiveScore += 1
    return positiveScore

# retrieve negative words from negative-words.txt file

negativeWords = []
with open('negative-words.txt', 'r', encoding='latin-1') as file2:
              data = file2.read().replace('\n', ' ')
negativeWords = preProcess(data)


# NEGATIVE SCORES

def calculateNegScore(filtered_tokens):
    negativeScore = 0
    for token in filtered_tokens:
        if(token in negativeWords):
            negativeScore += 1
            
    return negativeScore

# POLARITY SCORE

def calculatePolarityScore(positiveScore, negativeScore):
    polarityScore = (positiveScore - negativeScore)/((positiveScore + negativeScore) + 0.000001)
    return polarityScore


# SUBJECTIVITY SCORE

def calculateSubScore(positiveScore, negativeScore, filtered_tokens):
    wordsAfterCleaning = len(filtered_tokens)
    subjectivityScore = (positiveScore + negativeScore)/ ((wordsAfterCleaning) + 0.000001)
    return subjectivityScore

# #### READABILITY ANALYSIS

# install and import dependencies

# pip install spacy -U
# pip install textstat -U
# python -m spacy download en_core_web_sm

import spacy
from textstat.textstat import textstatistics

# Splits the text into sentences, using 
# Spacy's sentence segmentation which can 

def break_sentences(text):
    nlp = spacy.load('en_core_web_sm')
    doc = nlp(text)
    return list(doc.sents)

# Returns Number of Words in the text
def word_count(text):
    sentences = break_sentences(text)
    words = 0
    for sentence in sentences:
        words += len([token for token in sentence])
    return words


# Returns the number of sentences in the text
def sentence_count(text):
    sentences = break_sentences(text)
    return len(sentences)


# Returns average sentence length
def avg_sentence_length(text):
    words = word_count(text)
    sentences = sentence_count(text)
    if(sentences == 0):
        return 0
    average_sentence_length = float(words / sentences)
    return average_sentence_length


def avg_word_length(text):
    totalChars = len(text)
    if(totalChars==0):
        return 0
    return float(totalChars/word_count(text))


# Textstat is a python package, to calculate statistics from 
# text to determine readability, 
# complexity and grade level of a particular corpus.

def syllables_count(word):
    return textstatistics().syllable_count(word)


# Returns the average number of syllables per
# word in the text


def avg_syllables_per_word(text):
    syllable = syllables_count(text)
    words = word_count(text)
    if(words == 0):
        return 0
    ASPW = float(syllable) / float(words)
    return round(ASPW, 1)


def difficult_words(text):
     
    nlp = spacy.load('en_core_web_sm')
    doc = nlp(text)
    # Find all words in the text
    words = []
    sentences = break_sentences(text)
    for sentence in sentences:
        words += [str(token) for token in sentence]
 
    # difficult words are those with syllables > 2
    # easy_word_set is provide by Textstat as 
    # a list of common words
    diff_words_set = set()
     
    for word in words:
        syllable_count = syllables_count(word)
        if word not in nlp.Defaults.stop_words and syllable_count > 2:
            diff_words_set.add(word)
 
    return len(diff_words_set)


def per_complex_words(text):
    if(word_count(text) == 0):
        return 0
    per_diff_words = (difficult_words(text) / word_count(text) * 100) + 5
    return per_diff_words


def fog_index(text):
    per_diff_words = per_complex_words(text)
    grade = 0.4 * (avg_sentence_length(text) + per_diff_words)
    return grade


def clean_word_count(text):
    clean_words = filterContent(text)
    return len(clean_words)


# To calculate Personal Pronouns mentioned in the text, we use regex to find the counts of the words -
# “I,” “we,” “my,” “ours,” and “us”. 
# Special care is taken so that the country name US is not included in the list.
import re
def personal_pronouns(text):
    pronounRegex = re.compile(r'\b(I|we|my|ours|(?-i:us))\b',re.I)
    pronouns = pronounRegex.findall(text.lower())
    return len(pronouns)




# function to read a file content

def readFile(fileName):
    try:
        with open(fileName, 'r', encoding="utf-8") as file:
                  data = file.read().replace('\n', ' ')
        return data
    except:
        return ""



# ###  CODE TO WRITE THE CALCULATED SCORES IN OutputDataStructure.xlsx FILE FOR EVERY URL OF WEB PAGE


wrkbk2 = openpyxl.load_workbook("OutputDataStructure.xlsx")

sh = wrkbk2.active 

for i in range(2, sh.max_row+1): 
    #get the values of first column that are file name
    cell_obj = sh.cell(row=i, column=1)
    url_id = cell_obj.value
    cell_obj = sh.cell(row=i, column=2)
    url = cell_obj.value
    
    if(url_id):
        print(i)
        #retrieve file content
        fileName = url_id + ".txt"
        fileContent = readFile(fileName)

        #filter the content by removing punctuations and stop words
        filteredTokens = filterContent(fileContent)

        #sentiment analysis
        positiveScore = calculatePosScore(filteredTokens)
        negativeScore = calculateNegScore(filteredTokens)
        polarityScore = calculatePolarityScore(positiveScore, negativeScore)
        subjectivityScore = calculateSubScore(positiveScore, negativeScore, filteredTokens)
        
        #readability analysis
        avgSentLen = avg_sentence_length(fileContent)
        perComplexWords = per_complex_words(fileContent)
        fogIndex = fog_index(fileContent)
        avgWordsPerSen = avg_sentence_length(fileContent)
        complexWordCount = difficult_words(fileContent)
        wordCount = clean_word_count(fileContent)
        syllPerWord = avg_syllables_per_word(fileContent)
        personalProunounCount = personal_pronouns(fileContent)
        avgWordLen = avg_word_length(fileContent)
        

        c3 = sh.cell(row=i, column=3)
        c4 = sh.cell(row=i, column=4)
        c5 = sh.cell(row=i, column=5)
        c6 = sh.cell(row=i, column=6)
        c7 = sh.cell(row=i, column=7)
        c8 = sh.cell(row=i, column=8)
        c9 = sh.cell(row=i, column=9)
        c10 = sh.cell(row=i, column=10)
        c11 = sh.cell(row=i, column=11)
        c12 = sh.cell(row=i, column=12)
        c13 = sh.cell(row=i, column=13)
        c14 = sh.cell(row=i, column=14)
        c15 = sh.cell(row=i, column=15)

        c3.value = positiveScore
        c4.value = negativeScore
        c5.value = polarityScore
        c6.value = subjectivityScore
        c7.value = avgSentLen
        c8.value = perComplexWords
        c9.value = fogIndex
        c10.value = avgWordsPerSen
        c11.value = complexWordCount
        c12.value = wordCount
        c13.value = syllPerWord
        c14.value = personalProunounCount
        c15.value = avgWordLen

#Modify the path here according to location of OutputDataStructure.xlsx file in your system by copying its path and pasting here
wrkbk2.save("C:\\Users\\manra\\DataEngineeringProject\\OutputDataStructure.xlsx")
